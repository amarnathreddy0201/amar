"""
based on topic we will subscribe and store it into excel format
"""
from openpyxl.styles import PatternFill,Fill,Color
from openpyxl.styles import colors
#import openpyxl.styles
#D:\Amarnath\mqtt\subscriber.py
#import openpyxl
#import openpyxl.styles as sty
#from colorama import Fore, Back, Style
import random
import logging
from openpyxl import *
import paho.mqtt.client as mqtt
import json
import random
from datetime import datetime
import os
from os import path
import time
from openpyxl.styles import colors
from openpyxl.styles import Font,Color

broker = 'brokername'                                                                     #broker name
topic = "House/bedroom/light"  #on which topic we will subscrobe



def check(i,ws3,k):
    ws3.cell(row=1,column=1,value="S.NO")
    ws3.cell(row=1,column=2,value=k)
    ws3.cell(row=1,column=3,value="No of packets")
    ws3.cell(row=1,column=4,value="result")
    max_row=ws3.max_row+1
    
    if i>48 and i< 50:
        ws3.cell(row=max_row,column=1,value=max_row-1)
        ws3.cell(row=max_row,column=2,value=k)
        ws3.cell(row=max_row,column=3,value=5)
        ws3.cell(row=max_row, column=4).fill=PatternFill(start_color = "0000FF00", fill_type = 'solid')
        #ws3.cell(row=max_row,column=4).value="pass"#PatternFill(fill_type='solid',fgColor="0d5330")#.value="pass"#.colour="00FF0000"
        #ws3.cell(row=max_row,column=4).value="pass"
    else:
        ws3.cell(row=max_row,column=1,value=max_row-1)
        ws3.cell(row=max_row,column=2,value=k)
        ws3.cell(row=max_row,column=3,value=6)
        ws3.cell(row=max_row, column=4).fill=PatternFill(start_color = "00FF0000", fill_type = 'solid')
        #ws3.cell(row=max_row,column=4).value="fail"#fill=sty.PatternFill(fill_type='solid',fgColor="0d5330")#.value="fail"#.color="0000FF00"
        
    
"""
on_connect for connecting client to broker
"""
def on_connect(client, userdata, flags, rc):#client
    
    if rc == 0:
        
        print("Connected to MQTT Broker!",rc)
    else:
        
        print("Failed to connect, return code %d\n", rc)
    client.subscribe(topic)





"""
on messege for get msg based on subscribe
"""

def on_message(client, userdata, msg):
    s=msg.payload.decode()
    print("subscribe",s)
                       
                                                         
    r=0                                                                  
    json_data={}                                                        #dictionary for store th data
    try:
        print("try")
        json_data = json.loads(s)                                                       #json_load will convert the json string into json object
                            
        print("in below main reading json_data:",json_data)                             #json data printing
        #r=0                   
        p=path.abspath("mqtt")                                                      #mqtt.xlsx is exists or not
        if p:
            print("yes")
            wb1=load_workbook(filename="mqtt4.xlsx")                                      # work book loading
            ws=wb1["ANXMYRTHCPPM002"]                                           #onwork book which sheet(it must and should exist)

            #wb2=load_workbook(filename="mqtt1.xlsx")
            ws1=wb1["ANXMYRTHCPPM003"]

            #wb3=load_workbook(filename="mqtt2.xlsx")
            ws2=wb1["ANXMYRTHCPPM004"]
            ws3=wb1["RESULT"]
            """wb=Workbook()
            ws=wb.active
            ws.title="ANXMYRTHCPPM002"
            wb1=Workbook()
            ws1=wb1.active
            ws1.title="ANXMYRTHCPPM003"
            wb2=Workbook()
            ws2=wb2.active
            ws2.title="ANXMYRTHCPPM004"
            """
            if (json_data["MP"]["MN"] == "ANXMYRTHCPPM002"):
                
                ws.cell(1,1,"DT")                                                        #headings in a column
                ws.cell(1,2,"MI")
                ws.cell(1,3,"SD")
                ws.cell(1,4,"CS")
                ws.cell(1,5,"PQ")
                ws.cell(1,6,"MN")
                ws.cell(1,7,"DL")
                
                
                max_row=ws.max_row +1
                
                ws.cell(row=max_row,column=1,value=json_data["DT"])
                ws.cell(row=max_row,column=2,value=json_data["MI"])
                ws.cell(row=max_row,column=3,value=json_data["SD"])
                ws.cell(row=max_row,column=4,value=json_data["MP"]["CS"])
                ws.cell(row=max_row,column=5,value=json_data["MP"]["PQ"])
                ws.cell(row=max_row,column=6,value=json_data["MP"]["MN"])
                ws.cell(row=max_row,column=7,value=json_data["MP"]["DL"])
                
                
                for i,j in enumerate(json_data["MP"]["MM"]):
                    ws.cell(row=max_row,column=8+i).value=j
                
                """for i in range(len(a)):
                    ws.cell(row=max_row,column=8+i,value=a[i])"""

                
                print("if1")
            
            if(json_data["MP"]["MN"] == "ANXMYRTHCPPM003"):
                ws1.cell(1,1,"DT")                                                        #headings in a column
                ws1.cell(1,2,"MI")
                ws1.cell(1,3,"SD")
                ws1.cell(1,4,"CS")
                ws1.cell(1,5,"PQ")
                ws1.cell(1,6,"MN")
                ws1.cell(1,7,"DL")
                
                max_row=ws1.max_row+1 
                
                ws1.cell(row=max_row,column=1,value=json_data["DT"])
                ws1.cell(row=max_row,column=2,value=json_data["MI"])
                ws1.cell(row=max_row,column=3,value=json_data["SD"])
                ws1.cell(row=max_row,column=4,value=json_data["MP"]["CS"])
                ws1.cell(row=max_row,column=5,value=json_data["MP"]["PQ"])
                ws1.cell(row=max_row,column=6,value=json_data["MP"]["MN"])
                ws1.cell(row=max_row,column=7,value=json_data["MP"]["DL"])

                
                
                for i,j in enumerate(json_data["MP"]["MM"]):
                    ws1.cell(row=max_row,column=8+i).value=j
                """for i in range(len(l)):
                    ws1.cell(row=max_row,column=8+i,value=l[i])"""
                
                print("if 2")

            
            if(json_data["MP"]["MN"]== "ANXMYRTHCPPM004"):
                
                ws2.cell(1,1,"DT")
                ws2.cell(1,2,"MI")
                ws2.cell(1,3,"SD")
                ws2.cell(1,4,"CS")
                ws2.cell(1,5,"PQ")
                ws2.cell(1,6,"MN")
                ws2.cell(1,7,"DL")
                
            
                max_row=ws2.max_row + 1 
                ws2.cell(row=max_row,column=1,value=json_data["DT"])
                ws2.cell(row=max_row,column=2,value=json_data["MI"])
                ws2.cell(row=max_row,column=3,value=json_data["SD"])
                ws2.cell(row=max_row,column=4,value=json_data["MP"]["CS"])
                ws2.cell(row=max_row,column=5,value=json_data["MP"]["PQ"])
                ws2.cell(row=max_row,column=6,value=json_data["MP"]["MN"])
                ws2.cell(row=max_row,column=7,value=json_data["MP"]["DL"])

                
                
                for i,j in enumerate(json_data["MP"]["MM"]):
                    ws2.cell(row=max_row,column=8+i).value=j
                """for i in range(len(k)):
                    ws2.cell(row=max_row,column=8+i,value=k[i])"""
                
                
                print("if 3")
                   
            wb1.save("mqtt4.xlsx")                   
            
    except Exception as e:                                                              #exception for which ever not json data we will handle this
        
        print("Exception is",e)                                                         #if any exception through this we will which type of exception
        print("exception")
        logging.basicConfig(filename="mqtt_log",format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')  #log file for creating dumy file
                    
                                                                                        #logging.basicConfig(filename="dummy",format='%(asctime)s %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        with open("mqtt_log","a") as a:
            a.write(s) 
            logging.warning('is when this event was logged.')
        
                            

"""
main starting
"""


if __name__ == '__main__':
    
    
    client = mqtt.Client("p1")                                                          #client

    client.connect(broker)                                                              #connect the broker to mqtt

    client.on_connect= on_connect                                                       #client on_connect
    client.on_message= on_message                                                       #on_message we get topic message
    client.loop_forever()
    
